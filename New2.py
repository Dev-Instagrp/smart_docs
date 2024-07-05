from os.path import splitext
from typing import List, Sequence, Tuple

import pandas as pd
from google.cloud import documentai
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

import os
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = "cac-smart-docs-5130258bf4be.json"

def online_process(project_id: str, location: str, processor_id: str, file_path: str, mime_type: str) -> documentai.Document:
    """
    Processes a document using the Document AI Online Processing API.
    """
    opts = {"api_endpoint": f"{location}-documentai.googleapis.com"}

    # Instantiates a client
    documentai_client = documentai.DocumentProcessorServiceClient(client_options=opts)

    # The full resource name of the processor, e.g.:
    # projects/project-id/locations/location/processor/processor-id
    resource_name = documentai_client.processor_path(project_id, location, processor_id)

    # Read the file into memory
    with open(file_path, "rb") as image:
        image_content = image.read()

        # Load Binary Data into Document AI RawDocument Object
        raw_document = documentai.RawDocument(content=image_content, mime_type=mime_type)

        # Configure the process request
        request = documentai.ProcessRequest(name=resource_name, raw_document=raw_document)

        # Use the Document AI client to process the sample form
        result = documentai_client.process_document(request=request)
        return result.document

def get_table_data(rows: Sequence[documentai.Document.Page.Table.TableRow], text: str) -> Tuple[List[List[str]], List[List[float]]]:
    """
    Get text data and confidence scores from table rows.
    """
    all_values: List[List[str]] = []
    all_confidences: List[List[float]] = []
    for row in rows:
        current_row_values: List[str] = []
        current_row_confidences: List[float] = []
        for cell in row.cells:
            current_row_values.append(text_anchor_to_text(cell.layout.text_anchor, text))
            current_row_confidences.append(cell.layout.confidence)
        all_values.append(current_row_values)
        all_confidences.append(current_row_confidences)
    return all_values, all_confidences

def text_anchor_to_text(text_anchor: documentai.Document.TextAnchor, text: str) -> str:
    """
    Document AI identifies table data by their offsets in the entirety of the document's text. This function converts offsets to a string.
    """
    response = ""
    for segment in text_anchor.text_segments:
        start_index = int(segment.start_index)
        end_index = int(segment.end_index)
        response += text[start_index:end_index]
    return response.strip().replace("\n", " ")

def identify_inaccurate_entries(df):
    # Example logic for identifying inaccuracies
    # This should be replaced with your actual logic
    mask = df.applymap(lambda x: isinstance(x, str) and 'error' in x)
    return mask

def highlight_inaccurate_entries(df, inaccuracies, output_path):
    wb = Workbook()
    ws = wb.active

    # Write DataFrame to Excel
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Apply highlighting
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
        for j, cell in enumerate(row):
            if inaccuracies.iat[i, j]:
                cell.fill = red_fill

    # Save the workbook
    wb.save(output_path)

PROJECT_ID = "cac-smart-docs"
LOCATION = "us"  # Format is 'us' or 'eu'
PROCESSOR_ID = "373b6eea27dcab8"  # Create processor before running sample

# The local file in your current working directory
FILE_PATH = "./New Testing/0.7 Black Pen.pdf"
# Refer to https://cloud.google.com/document-ai/docs/file-types for supported file types
MIME_TYPE = "application/pdf"

document = online_process(project_id=PROJECT_ID, location=LOCATION, processor_id=PROCESSOR_ID, file_path=FILE_PATH, mime_type=MIME_TYPE)

header_row_values: List[List[str]] = []
header_row_confidences: List[List[float]] = []
body_row_values: List[List[str]] = []
body_row_confidences: List[List[float]] = []

# Input Filename without extension
output_file_prefix = splitext(FILE_PATH)[0]

for page in document.pages:
    for index, table in enumerate(page.tables):
        header_row_values, header_row_confidences = get_table_data(table.header_rows, document.text)
        body_row_values, body_row_confidences = get_table_data(table.body_rows, document.text)

        # Create a Pandas DataFrame to print the values in tabular format
        df = pd.DataFrame(data=body_row_values, columns=pd.MultiIndex.from_arrays(header_row_values))

        # Identify inaccuracies
        inaccuracies = identify_inaccurate_entries(df)

        # Output filename
        output_filename = f"{output_file_prefix}_pg{page.page_number}_tb{index}.xlsx"
        
        # Highlight inaccuracies and save to Excel
        highlight_inaccurate_entries(df, inaccuracies, output_filename)
